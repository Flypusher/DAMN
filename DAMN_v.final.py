"""
DAMN

Drosphila
Activity
Monitor (in excel)
Navigator


Warning: Does not work with :30 interval data since there are
alternating 0 rows"""

import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
import random


"""Importing DAM CSV to Excel"""
CSV_file = input("Type in CSV file name: ")
#Copy File URI, not name or path i.e. should like like file://D:"name".txt
excel_file = input("Type in what you want to name the excel sheet as: ")
excel_name = excel_file + ".xlsx"
df = pd.read_csv(CSV_file, delimiter='\t', header = None)
#makes csv file into a dataframe
# "\t" for tab separated data or "," for CSV for delimiter
###df.to_excel(excel_name, index = False, header = None)
#makes dataframe into excel with the excel name you made


"""Cutting out all data before the DAMs are warmed up i.e.  51 to 1"""
first_index = df[df.iloc[:, 3] == 1].index[0]
#iloc[:,3] = fourth column i.e. DAM warm up column D
df = df.iloc[first_index - 1:]
#only keeps data of row and below, with 0 as first row


"""Moves D and J // On/Off and Lights column to end"""
cols = df.columns.tolist()
columns_to_move = [cols[3], cols[9]] #i.e. columns D and E as index values
new_order = [col for col in cols if col not in columns_to_move] + columns_to_move
df = df[new_order]



"""deletes 5th column i.e. E and the 4 following ones including I"""
df.drop(columns=df.columns[3:8], inplace=True)


"""header"""
header = (
    ["Number", "Date", "Time"] + 
    [f'M{i}' for i in range(1, 33)] + 
    ["On/Off", "Light"])
df.columns = header
df.to_excel(excel_name, index=False, header = True)



#Time Selector
threshold = float(input("Track time for activity greater than: "))

wb = load_workbook(excel_name)
ws = wb.active

if "Threshold Matches (Change Name)" in wb.sheetnames:
    del wb["Threshold Matches (Change Name)"]
out_ws = wb.create_sheet("Threshold Matches (Change Name)")

out_ws.append([f"M{i}" for i in range(1,33)])

results = []

NUM_COLUMNS = 32
m_start_col = 4
m_end_col = m_start_col + NUM_COLUMNS - 1

for i in range (m_start_col, m_end_col + 1): #problem with M1 not being tracked
    count = 0
    values = []
    
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column = i)
        try:
            if float(cell.value) > threshold:
                number_val = ws.cell(row=row, column=1).value
                values.append(number_val)
                """count += 1
                if count == threshold:
                    break"""
                if len(values) == 2:
                    break
        except (TypeError, ValueError):
            continue
    
    #Pad with None if fewer than 2 values found
    while len(values) < 2:
        values.append(None)
    
    results.append(values)
        
for row_idx in range(2):
    row = [results[col_idx][row_idx] for col_idx in range(NUM_COLUMNS)]
    out_ws.append(row)





#relative time table
base_val = ws["A3"].value                     # the reference value in A3

mark_col = 1
start_row = out_ws.max_row + 1
out_ws.cell(row=start_row, column=mark_col, value= "Start Time (A3)")
out_ws.cell(row=start_row +1, column=mark_col, value=base_val)

blank_row_idx = out_ws.max_row + 1               # leave a blank line
out_ws.append([])                                # visual spacer

# header for the difference table
delta_header = [f"ΔM{i}" for i in range(1, 33)]
out_ws.append(delta_header)

# the two original data rows live at rows 2 and 3 in out_ws
for src_excel_row in (2, 3):
    delta_row = []
    for col in range(1, 33):                     # columns A → AF (M1‑M32)
        cell_val = out_ws.cell(src_excel_row, col).value
        if cell_val is not None and base_val is not None:
            delta_row.append(cell_val - base_val)  # time – A3
        else:
            delta_row.append(None)                 # keep blanks where no data
    out_ws.append(delta_row)





wb.save(excel_name)









































































chance = random.random()
if chance < .1:
    print("""\n
          \n
    For God so loved the world, \n
      that he gave his only begotten Son, \n
      that whosoever believeth in him \n
      should not perish, \n
      but have everlasting life.
      \n - John 3:16""")